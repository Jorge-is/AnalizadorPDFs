"""
export_controller.py
Controlador de exportación a Excel.

Recibe por POST:
  - columnas (list[str]): nombres de las columnas a incluir.
  - filas    (list[int], opcional): índices (base-0) de las filas a exportar.
                                    Si está vacío, exporta TODAS.

Genera un .xlsx estilizado y devuelve la URL de descarga.
"""
from flask import (
    Blueprint, request, jsonify, url_for,
    send_from_directory, current_app
)
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
import os
from models import db, Article, Collection
from flask_login import login_required, current_user

export_bp = Blueprint('export', __name__)

# Mapeo frontend-key → clave real en el dict de metadata
COLUMNAS_VALIDAS = {
    "titulo":           "titulo",
    "autores":          "autores",
    "anio":             "anio",
    "tema":             "tema",
    "pais":             "pais",
    "palabras":         "palabras",
    "resumen":          "resumen",
    "paginas_imagenes": "paginas_imagenes",
}

# Etiquetas legibles para la cabecera del Excel
ETIQUETAS = {
    "titulo":           "Título",
    "autores":          "Autores",
    "anio":             "Año",
    "tema":             "Tema",
    "pais":             "País",
    "palabras":         "Palabras Clave",
    "resumen":          "Resumen",
    "paginas_imagenes": "Págs / Imgs",
}

# Estilos Excel
_FILL_HEADER = PatternFill(start_color="0D1421", end_color="0D1421", fill_type="solid")
_FONT_HEADER = Font(color="3B7DE8", bold=True, name="Calibri", size=11)
_FONT_BODY   = Font(color="E8EDF5", name="Calibri", size=10)
_BORDER      = Border(
    left=Side(style='thin', color="1E2D47"),
    right=Side(style='thin', color="1E2D47"),
    top=Side(style='thin', color="1E2D47"),
    bottom=Side(style='thin', color="1E2D47"),
)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
_ALIGN_TOP    = Alignment(horizontal="left", vertical="top", wrap_text=True)


@export_bp.route("/exportar_excel", methods=["POST"])
@login_required
def exportar():
    # ── 1. Obtener datos de la BD ─────────────────────────────────
    col = Collection.query.filter_by(user_id=current_user.id, name="Default").first()
    if not col:
        return jsonify(error="No hay datos para exportar."), 400

    todos_los_articulos = Article.query.filter_by(
        collection_id=col.id, status='completed'
    ).all()

    if not todos_los_articulos:
        return jsonify(error="No hay artículos procesados para exportar."), 400

    todos_los_datos = [a.get_metadata() for a in todos_los_articulos]

    # ── 2. Validar columnas ───────────────────────────────────────
    columnas_raw = request.form.getlist('columnas')
    # Filtrar solo columnas conocidas y asegurar que 'titulo' siempre esté
    columnas = [c for c in columnas_raw if c in COLUMNAS_VALIDAS]
    if 'titulo' not in columnas:
        columnas.insert(0, 'titulo')

    if len(columnas) == 0:
        return jsonify(error="No se seleccionaron columnas válidas."), 400

    # ── 3. Filtrar filas ──────────────────────────────────────────
    filas_raw = request.form.getlist('filas')
    if filas_raw:
        try:
            indices = [int(i) for i in filas_raw if i.strip().isdigit()]
            # Filtrar índices fuera de rango
            datos_exportar = [
                todos_los_datos[i] for i in indices
                if 0 <= i < len(todos_los_datos)
            ]
        except (ValueError, IndexError):
            datos_exportar = todos_los_datos
    else:
        datos_exportar = todos_los_datos   # Exportar todo si no hay selección

    if not datos_exportar:
        return jsonify(error="No hay filas válidas para exportar."), 400

    # ── 4. Crear Excel ────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "RSL Datos"

    # Cabecera
    encabezado = [ETIQUETAS.get(col_key, col_key.capitalize()) for col_key in columnas]
    ws.append(encabezado)

    for col_num, _ in enumerate(encabezado, start=1):
        celda = ws.cell(row=1, column=col_num)
        celda.fill      = _FILL_HEADER
        celda.font      = _FONT_HEADER
        celda.border    = _BORDER
        celda.alignment = _ALIGN_CENTER

    # Filas de datos
    fill_par   = PatternFill(start_color="0D1421", end_color="0D1421", fill_type="solid")
    fill_impar = PatternFill(start_color="131C2E", end_color="131C2E", fill_type="solid")

    for fila_num, dato in enumerate(datos_exportar, start=1):
        fila = [dato.get(COLUMNAS_VALIDAS[col_key], "") for col_key in columnas]
        ws.append(fila)

        fill = fill_impar if fila_num % 2 == 0 else fill_par
        for col_num in range(1, len(fila) + 1):
            celda = ws.cell(row=fila_num + 1, column=col_num)
            celda.fill      = fill
            celda.font      = _FONT_BODY
            celda.border    = _BORDER
            celda.alignment = _ALIGN_TOP

    # Ancho automático de columnas
    for col_idx, col_key in enumerate(columnas, start=1):
        max_len = max(
            (len(str(dato.get(COLUMNAS_VALIDAS[col_key], "") or "")) for dato in datos_exportar),
            default=10,
        )
        max_len = min(max(max_len, len(encabezado[col_idx - 1])) + 4, 60)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len

    # Fijar fila de encabezado
    ws.freeze_panes = "A2"

    # ── 5. Guardar y devolver URL ─────────────────────────────────
    nombre_archivo = "rsl_datos.xlsx"
    ruta_archivo   = os.path.join(current_app.config['UPLOAD_FOLDER'], nombre_archivo)
    wb.save(ruta_archivo)

    return jsonify(url=url_for("export.descargar_excel"))


@export_bp.route("/descargar_excel")
@login_required
def descargar_excel():
    return send_from_directory(
        current_app.config['UPLOAD_FOLDER'],
        "rsl_datos.xlsx",
        as_attachment=True,
        download_name="rsl_datos.xlsx",
    )
